#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SQL编辑器组件
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableView, 
    QPlainTextEdit, QTabWidget, QLabel, QSplitter, QFileDialog, QMenu, QMessageBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QAbstractTableModel
from PyQt6.QtGui import QSyntaxHighlighter, QTextCharFormat, QColor, QFont, QAction
import re
import csv
import os
from openpyxl import Workbook


class SQLEditorWidget(QWidget):
    """SQL编辑器组件"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self._init_ui()
        self._init_highlighter()
        self.connection_info = None
    
    def _init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout()
        
        # 工具栏
        toolbar_layout = QHBoxLayout()
        
        self.run_button = QPushButton("运行")
        self.run_button.clicked.connect(self._run_sql)
        toolbar_layout.addWidget(self.run_button)
        
        self.stop_button = QPushButton("停止")
        self.stop_button.clicked.connect(self._stop_sql)
        self.stop_button.setEnabled(False)
        toolbar_layout.addWidget(self.stop_button)
        
        # 导出按钮
        self.export_button = QPushButton("导出")
        self.export_button.clicked.connect(self._show_export_menu)
        toolbar_layout.addWidget(self.export_button)
        
        toolbar_layout.addStretch()
        
        layout.addLayout(toolbar_layout)
        
        # 导出菜单
        self.export_menu = QMenu()
        self.export_csv_action = QAction("导出为CSV", self)
        self.export_csv_action.triggered.connect(lambda: self._export("csv"))
        self.export_menu.addAction(self.export_csv_action)
        
        self.export_excel_action = QAction("导出为Excel", self)
        self.export_excel_action.triggered.connect(lambda: self._export("excel"))
        self.export_menu.addAction(self.export_excel_action)
        
        # 编辑器和结果区域
        splitter = QSplitter(Qt.Orientation.Vertical)
        
        # SQL编辑器
        self.editor = QPlainTextEdit()
        self.editor.setPlaceholderText("在此输入SQL语句...")
        self.editor.setFont(QFont("Consolas", 10))
        splitter.addWidget(self.editor)
        
        # 结果标签页
        self.result_tabs = QTabWidget()
        splitter.addWidget(self.result_tabs)
        
        # 设置分割比例
        splitter.setSizes([400, 300])
        
        layout.addWidget(splitter)
        
        self.setLayout(layout)
        
        # 快捷键
        self.editor.setShortcutEnabled(True)
    
    def _init_highlighter(self):
        """初始化语法高亮"""
        self.highlighter = SQLHighlighter(self.editor.document())
    
    def _run_sql(self):
        """运行SQL"""
        # 获取选中的SQL或全部SQL
        selected_text = self.editor.textCursor().selectedText()
        if selected_text:
            sql = selected_text
        else:
            sql = self.editor.toPlainText()
        
        if not sql.strip():
            return
        
        # 禁用按钮
        self.run_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        
        # 创建执行线程
        self.worker = SQLWorker(sql, self.connection_info)
        self.worker.result_ready.connect(self._on_result_ready)
        self.worker.error_occurred.connect(self._on_error_occurred)
        self.worker.finished.connect(self._on_worker_finished)
        self.worker.start()
    
    def _stop_sql(self):
        """停止SQL执行"""
        if hasattr(self, 'worker') and self.worker.isRunning():
            self.worker.terminate()
            self._on_worker_finished()
    
    def _on_result_ready(self, results):
        """处理执行结果"""
        # 清空现有标签页
        while self.result_tabs.count() > 0:
            self.result_tabs.removeTab(0)
        
        # 添加结果标签页
        for i, result in enumerate(results):
            if isinstance(result, list) and result:
                # 创建结果视图
                result_widget = QWidget()
                result_layout = QVBoxLayout()
                
                # 结果表格
                table_view = QTableView()
                result_layout.addWidget(table_view)
                
                # 创建模型并设置数据
                model = SQLResultModel(result)
                table_view.setModel(model)
                
                # 调整列宽
                table_view.resizeColumnsToContents()
                
                result_widget.setLayout(result_layout)
                self.result_tabs.addTab(result_widget, f"结果集 {i+1}")
            else:
                # 非查询结果
                result_widget = QWidget()
                result_layout = QVBoxLayout()
                
                label = QLabel(f"执行成功，影响行数: {result}")
                result_layout.addWidget(label)
                
                result_widget.setLayout(result_layout)
                self.result_tabs.addTab(result_widget, f"结果 {i+1}")
    
    def _on_error_occurred(self, error):
        """处理错误"""
        # 清空现有标签页
        while self.result_tabs.count() > 0:
            self.result_tabs.removeTab(0)
        
        # 添加错误标签页
        error_widget = QWidget()
        error_layout = QVBoxLayout()
        
        label = QLabel(f"执行错误: {error}")
        label.setStyleSheet("color: red")
        error_layout.addWidget(label)
        
        error_widget.setLayout(error_layout)
        self.result_tabs.addTab(error_widget, "错误")
    
    def _on_worker_finished(self):
        """工作线程结束"""
        self.run_button.setEnabled(True)
        self.stop_button.setEnabled(False)
    
    def _show_export_menu(self):
        """显示导出菜单"""
        self.export_menu.exec(self.export_button.mapToGlobal(self.export_button.rect().bottomLeft()))
    
    def _export(self, file_type):
        """导出数据"""
        current_tab = self.result_tabs.currentWidget()
        if not current_tab:
            return
        
        # 查找表格视图
        table_view = None
        for widget in current_tab.findChildren(QTableView):
            table_view = widget
            break
        
        if not table_view:
            return
        
        model = table_view.model()
        if not model:
            return
        
        # 获取数据
        data = []
        headers = []
        
        # 获取表头
        for col in range(model.columnCount()):
            header = model.headerData(col, Qt.Orientation.Horizontal)
            headers.append(header)
        
        # 获取数据
        for row in range(model.rowCount()):
            row_data = []
            for col in range(model.columnCount()):
                index = model.index(row, col)
                value = model.data(index)
                row_data.append(value)
            data.append(row_data)
        
        if not data:
            return
        
        # 选择保存路径
        if file_type == "csv":
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存CSV文件", "", "CSV Files (*.csv)"
            )
            if file_path:
                self._export_to_csv(file_path, headers, data)
        elif file_type == "excel":
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存Excel文件", "", "Excel Files (*.xlsx)"
            )
            if file_path:
                self._export_to_excel(file_path, headers, data)
    
    def _export_to_csv(self, file_path, headers, data):
        """导出为CSV文件"""
        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(data)
            QMessageBox.information(self, "成功", "数据已导出为CSV文件")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def _export_to_excel(self, file_path, headers, data):
        """导出为Excel文件"""
        try:
            wb = Workbook()
            ws = wb.active
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, row_data in enumerate(data, 2):
                for col, value in enumerate(row_data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            wb.save(file_path)
            QMessageBox.information(self, "成功", "数据已导出为Excel文件")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def set_connection_info(self, connection_info):
        """设置数据库连接信息"""
        self.connection_info = connection_info


class SQLHighlighter(QSyntaxHighlighter):
    """SQL语法高亮"""
    
    def __init__(self, document):
        super().__init__(document)
        self._init_formats()
        self._init_rules()
    
    def _init_formats(self):
        """初始化格式"""
        # 关键字
        self.keyword_format = QTextCharFormat()
        self.keyword_format.setForeground(QColor(120, 120, 255))
        self.keyword_format.setFontWeight(QFont.Weight.Bold)
        
        # 字符串
        self.string_format = QTextCharFormat()
        self.string_format.setForeground(QColor(255, 120, 120))
        
        # 注释
        self.comment_format = QTextCharFormat()
        self.comment_format.setForeground(QColor(120, 255, 120))
        
        # 数字
        self.number_format = QTextCharFormat()
        self.number_format.setForeground(QColor(255, 255, 120))
    
    def _init_rules(self):
        """初始化规则"""
        # SQL关键字
        keywords = [
            'SELECT', 'FROM', 'WHERE', 'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'DROP',
            'ALTER', 'TABLE', 'DATABASE', 'VIEW', 'INDEX', 'TRIGGER', 'PROCEDURE',
            'FUNCTION', 'BEGIN', 'END', 'IF', 'ELSE', 'WHILE', 'FOR', 'IN', 'JOIN',
            'LEFT', 'RIGHT', 'INNER', 'OUTER', 'ON', 'AS', 'GROUP', 'BY', 'ORDER',
            'HAVING', 'LIMIT', 'OFFSET', 'AND', 'OR', 'NOT', 'LIKE', 'IN', 'BETWEEN',
            'IS', 'NULL', 'TRUE', 'FALSE'
        ]
        
        # 构建规则
        self.rules = []
        
        # 关键字规则
        for keyword in keywords:
            pattern = re.compile(r'\\b' + keyword + r'\\b', re.IGNORECASE)
            self.rules.append((pattern, self.keyword_format))
        
        # 字符串规则
        self.rules.append((re.compile(r'"[^"]*"'), self.string_format))
        self.rules.append((re.compile(r"'[^']*'"), self.string_format))
        
        # 注释规则
        self.rules.append((re.compile(r'--.*$'), self.comment_format))
        self.rules.append((re.compile(r'#.*$'), self.comment_format))
        
        # 数字规则
        self.rules.append((re.compile(r'\\b\d+\\b'), self.number_format))
    
    def highlightBlock(self, text):
        """高亮文本块"""
        for pattern, format in self.rules:
            for match in pattern.finditer(text):
                self.setFormat(match.start(), match.end() - match.start(), format)


class SQLWorker(QThread):
    """SQL执行工作线程"""
    
    result_ready = pyqtSignal(list)
    error_occurred = pyqtSignal(str)
    
    def __init__(self, sql, connection_info=None):
        super().__init__()
        self.sql = sql
        self.connection_info = connection_info
    
    def run(self):
        """运行"""
        try:
            results = []
            
            if self.connection_info:
                # 使用数据库驱动执行SQL
                from app.services.driver_factory import DriverFactory
                
                # 创建驱动
                driver = DriverFactory.create_driver(self.connection_info["type"])
                
                # 连接数据库
                success = driver.connect(self.connection_info)
                if success:
                    # 执行SQL
                    result = driver.execute(self.sql)
                    
                    # 处理结果
                    if isinstance(result, list):
                        # 查询结果
                        results.append(result)
                    else:
                        # 非查询结果（影响行数）
                        results.append(result)
                    
                    # 断开连接
                    driver.disconnect()
                else:
                    self.error_occurred.emit("连接数据库失败")
                    return
            else:
                # 没有连接信息，报错
                self.error_occurred.emit("数据库连接信息未设置")
                return
            
            self.result_ready.emit(results)
        except Exception as e:
            self.error_occurred.emit(str(e))


class SQLResultModel(QAbstractTableModel):
    """SQL结果模型"""
    
    def __init__(self, data):
        super().__init__()
        self.data = data
        if data:
            self.columns = list(data[0].keys())
        else:
            self.columns = []
    
    def rowCount(self, parent=None):
        return len(self.data)
    
    def columnCount(self, parent=None):
        return len(self.columns)
    
    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            row = index.row()
            col = index.column()
            if row < len(self.data) and col < len(self.columns):
                return str(self.data[row][self.columns[col]])
        return None
    
    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                if section < len(self.columns):
                    return self.columns[section]
        return None
